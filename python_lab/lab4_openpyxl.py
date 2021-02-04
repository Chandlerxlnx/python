# -*- coding: utf-8 -*-
"""
Created on Sun Jan 24 21:32:28 2021

@author: Zandra
"""

#例9-12   把记事本文件test.txt转换成Excel 2007+文件。假设test.txt文件中第一行为表头，从第二行开始是实际数据，并且表头和数据行中的不同字段信息都是用逗号分隔。
from openpyxl import Workbook
from random import randint

def main(txtFileName):
    new_XlsxFileName = txtFileName[:-3] + 'xlsx'
    wb = Workbook()
    ws = wb.worksheets[0]
    with open(txtFileName) as fp:
        for line in fp:
            line = line.strip().split(',')
            line =[(int(s0) if s0.isnumeric() 
                 else float(s0) if s0.replace('.','',1).isnumeric() 
                    else  s0) for s0 in line]
            ws.append(line)
    wb.save(new_XlsxFileName)

# create test.txt file
line_format ='{0:d},{1:3.3f}\n'
with open(r'labfile\test.txt','w') as fp:
    fp.writelines("year,revenue\n")
    for n in range(10):
        fp.write(line_format.format((2000+n), randint(100,1000)/10))
    
main(r'labfile\test.txt')


#例9-13  输出Excel文件中单元格中公式计算结果。

import openpyxl

#打开Excel文件
wb = openpyxl.load_workbook(r'labfile/data.xlsx', data_only=True)

#获取WorkSheet
ws = wb.worksheets[1]

#遍历Excel文件所有行，假设下标为2的列中是公式
for row in ws.rows:
    print(row[2].value)
    #print(row)

print(ws['C2'])
print(wb.worksheets[1].cell(3,3).internal_value)
